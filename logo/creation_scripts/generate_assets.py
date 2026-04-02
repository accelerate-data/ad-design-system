#!/usr/bin/env python3
from __future__ import annotations

import argparse
import io
import re
import shutil
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Iterable

import cairosvg
from openpyxl import load_workbook
from PIL import Image


REQUIRED_COLUMNS = [
    "Platform/Use",
    "Logo Type",
    "Logo Color",
    "Format",
    "Size Spec",
    "Export Px (W×H)",
    "Suggested Path",
    "Filename",
    "Notes",
]

REQUIRED_SOURCE_MASTERS = [
    "icon-light.svg",
    "icon-dark.svg",
    "icon-monochrome.svg",
    "logo-light.svg",
    "logo-dark.svg",
    "logo-icon-lockup-light.svg",
    "logo-icon-lockup-dark.svg",
]


@dataclass
class RowRecord:
    row_number: int
    platform_use: str
    logo_type: str
    logo_color: str
    fmt: str
    size_spec: str
    export_px: str
    suggested_path: str
    filename: str
    notes: str


@dataclass
class PlannedOutput:
    row_number: int
    platform_use: str
    logo_type: str
    logo_color: str
    fmt: str
    size_spec: str
    export_px_for_manifest: str
    suggested_path: str
    final_filename: str
    relative_output_path: PurePosixPath
    source_svg: Path
    png_width: int | None = None
    png_height: int | None = None
    ico_sizes: list[int] | None = None


_HERE = Path(__file__).parent


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate logo exports from checklist XLSX.")
    parser.add_argument("--output", default=str(_HERE.parent / "output_from_scripts"), help="Output directory (default: output_from_scripts/ next to creation_scripts/)")
    parser.add_argument("--sources", default=str(_HERE.parent / "archive" / "svg_sources"), help="Sources directory (default: archive/svg_sources/)")
    parser.add_argument(
        "--excel",
        default=str(_HERE / "logo-export-checklist.xlsx"),
        help="Path to export checklist workbook (default: logo-export-checklist.xlsx)",
    )
    parser.add_argument(
        "--sheet",
        default="Export Checklist",
        help='Sheet name to read (default: "Export Checklist")',
    )
    parser.add_argument("--dry-run", action="store_true", help="Do not write files, only print actions.")
    return parser.parse_args()


def as_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    return text.strip()


def canonical_export_px(value: str) -> str:
    text = as_text(value).lower()
    text = text.replace("×", "x")
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"\s+", "", text)
    return text


def normalize_logo_type(value: str) -> str:
    normalized = as_text(value).lower().replace(" ", "")
    if normalized == "logo+icon":
        return "logo+icon"
    if normalized == "logooricon":
        return "logo or icon"
    if normalized == "icon":
        return "icon"
    if normalized == "logo":
        return "logo"
    return as_text(value).lower()


def normalize_color(value: str) -> str:
    return as_text(value).lower()


def normalize_format(value: str) -> str:
    return as_text(value).lower()


def row_summary(row: RowRecord, filename: str | None = None) -> str:
    chosen_filename = filename if filename is not None else row.filename
    return (
        f'row={row.row_number} platform="{row.platform_use}" '
        f'filename="{chosen_filename}" format="{row.fmt}"'
    )


def ensure_required_sources(sources_dir: Path) -> None:
    missing = [name for name in REQUIRED_SOURCE_MASTERS if not (sources_dir / name).exists()]
    if missing:
        joined = ", ".join(missing)
        raise RuntimeError(f"Missing required source file(s): {joined}")


def build_column_index(header_row: Iterable[object]) -> dict[str, int]:
    values = [as_text(v) for v in header_row]
    index: dict[str, int] = {}
    for i, value in enumerate(values):
        if value:
            index[value] = i
    missing = [col for col in REQUIRED_COLUMNS if col not in index]
    if missing:
        raise RuntimeError(f"Missing required column(s): {', '.join(missing)}")
    return index


def read_rows(excel_path: Path, sheet_name: str) -> list[RowRecord]:
    workbook = load_workbook(excel_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(
            f'Sheet "{sheet_name}" not found in {excel_path.name}. '
            f"Available sheets: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(min_row=1, values_only=True)
    header = next(iterator, None)
    if header is None:
        raise RuntimeError(f'Sheet "{sheet_name}" is empty.')

    col = build_column_index(header)
    rows: list[RowRecord] = []

    for excel_row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        line = list(values)
        if all(as_text(v) == "" for v in line):
            continue

        def cell(name: str) -> str:
            idx = col[name]
            if idx >= len(line):
                return ""
            return as_text(line[idx])

        rows.append(
            RowRecord(
                row_number=excel_row_number,
                platform_use=cell("Platform/Use"),
                logo_type=cell("Logo Type"),
                logo_color=cell("Logo Color"),
                fmt=cell("Format"),
                size_spec=cell("Size Spec"),
                export_px=cell("Export Px (W×H)"),
                suggested_path=cell("Suggested Path"),
                filename=cell("Filename"),
                notes=cell("Notes"),
            )
        )

    return rows


def apply_color_suffix(filename: str, color: str) -> str:
    lower = filename.lower()
    if "-light" in lower or "-dark" in lower:
        return filename
    path = PurePosixPath(filename)
    return str(path.with_name(f"{path.stem}-{color}{path.suffix}"))


def insert_suffix(filename: str, suffix: str) -> str:
    path = PurePosixPath(filename)
    return str(path.with_name(f"{path.stem}-{suffix}{path.suffix}"))


def normalize_suggested_path(value: str) -> tuple[PurePosixPath, str]:
    text = as_text(value).replace("\\", "/")
    while text.startswith("./"):
        text = text[2:]
    text = text.strip()

    if text == "":
        return PurePosixPath("."), ""

    if text.startswith("/"):
        raise ValueError("Suggested Path must be relative (absolute path provided).")
    if re.match(r"^[a-zA-Z]:[\\/]", text):
        raise ValueError("Suggested Path must be relative (Windows absolute path provided).")

    path = PurePosixPath(text)
    if any(part == ".." for part in path.parts):
        raise ValueError("Suggested Path must not contain '..'.")

    normalized = path.as_posix().rstrip("/")
    if normalized in ("", "."):
        return PurePosixPath("."), ""

    return PurePosixPath(normalized), f"{normalized}/"


def parse_png_range(export_px: str) -> list[int] | None:
    spec = canonical_export_px(export_px)
    match = re.fullmatch(r"(\d+)-(\d+)x(\d+)-(\d+)", spec)
    if not match:
        return None
    a, b, c, d = [int(part) for part in match.groups()]
    if a == c and b == d and a < b:
        return [a, b]
    return None


def parse_numeric_attr(value: str | None) -> float | None:
    if not value:
        return None
    match = re.match(r"^\s*([0-9]+(?:\.[0-9]+)?)", value)
    if not match:
        return None
    return float(match.group(1))


def get_svg_aspect_ratio(svg_path: Path, ratio_cache: dict[Path, float]) -> float:
    if svg_path in ratio_cache:
        return ratio_cache[svg_path]

    try:
        root = ET.parse(svg_path).getroot()
    except ET.ParseError as exc:
        raise ValueError(f"Unable to parse SVG for aspect ratio: {svg_path}") from exc

    view_box = root.attrib.get("viewBox")
    if view_box:
        parts = re.split(r"[,\s]+", view_box.strip())
        if len(parts) == 4:
            width = float(parts[2])
            height = float(parts[3])
            if width > 0 and height > 0:
                ratio = width / height
                ratio_cache[svg_path] = ratio
                return ratio

    width_attr = parse_numeric_attr(root.attrib.get("width"))
    height_attr = parse_numeric_attr(root.attrib.get("height"))
    if width_attr and height_attr and height_attr > 0:
        ratio = width_attr / height_attr
        ratio_cache[svg_path] = ratio
        return ratio

    raise ValueError(f"Cannot determine aspect ratio from {svg_path}.")


def parse_png_dimensions(
    export_px: str, source_svg: Path, ratio_cache: dict[Path, float]
) -> tuple[int, int]:
    spec = canonical_export_px(export_px)

    fixed = re.fullmatch(r"(\d+)x(\d+)", spec)
    if fixed:
        width, height = int(fixed.group(1)), int(fixed.group(2))
        return width, height

    auto_height = re.fullmatch(r"~\(auto\)x(\d+)", spec)
    if auto_height:
        height = int(auto_height.group(1))
        ratio = get_svg_aspect_ratio(source_svg, ratio_cache)
        width = max(1, int(round(ratio * height)))
        return width, height

    raise ValueError(f'Unsupported PNG Export Px pattern "{export_px}".')


def parse_ico_sizes(export_px: str) -> list[int]:
    raw = as_text(export_px)
    if raw == "":
        raise ValueError("ICO Export Px is empty.")

    sizes: list[int] = []
    for token in raw.split(","):
        part = token.strip().lower().replace(" ", "")
        if part == "":
            continue
        match = re.fullmatch(r"(\d+)(?:x\d+)?", part)
        if not match:
            raise ValueError(f'Invalid ICO size token "{token}".')
        size = int(match.group(1))
        if size <= 0:
            raise ValueError(f'Invalid ICO size "{token}".')
        if size not in sizes:
            sizes.append(size)

    if not sizes:
        raise ValueError("No ICO sizes were parsed.")
    return sizes


def resolve_effective_logo_type(row: RowRecord, filename: str) -> str:
    logo_type = normalize_logo_type(row.logo_type)
    if logo_type != "logo or icon":
        return logo_type

    platform_lower = row.platform_use.lower()
    filename_lower = filename.lower()
    if "logo" in filename_lower or "docs" in platform_lower or "email" in platform_lower:
        return "logo"
    return "icon"


def resolve_source_svg(
    row: RowRecord,
    color: str,
    filename: str,
    fmt: str,
    sources_dir: Path,
) -> tuple[Path | None, str | None]:
    if fmt == "svg" and filename.lower() == "safari-pinned-tab.svg":
        return sources_dir / "icon-monochrome.svg", None

    effective_type = resolve_effective_logo_type(row, filename)
    mapping = {
        ("icon", "light"): "icon-light.svg",
        ("icon", "dark"): "icon-dark.svg",
        ("icon", "monochrome"): "icon-monochrome.svg",
        ("logo", "light"): "logo-light.svg",
        ("logo", "dark"): "logo-dark.svg",
        ("logo+icon", "light"): "logo-icon-lockup-light.svg",
        ("logo+icon", "dark"): "logo-icon-lockup-dark.svg",
    }

    source_name = mapping.get((effective_type, color))
    if not source_name:
        return None, f'No source mapping for logo_type="{row.logo_type}" color="{color}".'
    return sources_dir / source_name, None


def reserve_relative_path(rel_path: PurePosixPath, used_paths: set[str]) -> PurePosixPath:
    parent = rel_path.parent
    stem = rel_path.stem
    suffix = rel_path.suffix

    candidate = rel_path
    count = 2
    while candidate.as_posix().lower() in used_paths:
        candidate = parent / f"{stem}-{count}{suffix}"
        count += 1

    used_paths.add(candidate.as_posix().lower())
    return candidate


def render_ico(source_svg: Path, destination: Path, sizes: list[int]) -> None:
    rendered: list[Image.Image] = []
    try:
        for size in sizes:
            png_bytes = cairosvg.svg2png(
                url=str(source_svg),
                output_width=size,
                output_height=size,
            )
            image = Image.open(io.BytesIO(png_bytes)).convert("RGBA")
            rendered.append(image)

        save_sizes = [(size, size) for size in sizes]
        base = rendered[0]
        try:
            base.save(
                destination,
                format="ICO",
                save_all=len(rendered) > 1,
                append_images=rendered[1:],
                sizes=save_sizes,
            )
        except Exception:
            # Fallback for Pillow builds that do not support append_images for ICO.
            base.save(destination, format="ICO", sizes=save_sizes)
    finally:
        for image in rendered:
            image.close()


def md_escape(value: str) -> str:
    return as_text(value).replace("|", "\\|").replace("\n", " ")


def write_manifest(
    manifest_path: Path,
    outputs: list[PlannedOutput],
    excel_filename: str,
    sheet_name: str,
) -> None:
    generated_at = datetime.now(timezone.utc).isoformat()

    lines = [
        "# Assets Manifest",
        "",
        f"- Generated At (UTC): `{generated_at}`",
        f"- Source Spreadsheet: `{excel_filename}`",
        f"- Sheet: `{sheet_name}`",
        "",
        "| Platform/Use | Logo Type | Logo Color | Format | Size Spec | Export Px (W×H) | Suggested Path | Filename | Relative Output Path |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]

    sorted_outputs = sorted(outputs, key=lambda item: (item.suggested_path, item.final_filename))
    for item in sorted_outputs:
        lines.append(
            "| "
            + " | ".join(
                [
                    md_escape(item.platform_use),
                    md_escape(item.logo_type),
                    md_escape(item.logo_color),
                    md_escape(item.fmt.upper()),
                    md_escape(item.size_spec),
                    md_escape(item.export_px_for_manifest),
                    md_escape(item.suggested_path),
                    md_escape(item.final_filename),
                    md_escape(item.relative_output_path.as_posix()),
                ]
            )
            + " |"
        )

    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def generate_outputs(args: argparse.Namespace) -> int:
    output_root = Path(args.output)
    sources_dir = Path(args.sources)
    excel_path = Path(args.excel)
    sheet_name = args.sheet
    dry_run = args.dry_run

    if not excel_path.exists():
        raise RuntimeError(f"Spreadsheet not found: {excel_path}")

    ensure_required_sources(sources_dir)
    rows = read_rows(excel_path, sheet_name)

    used_relative_paths: set[str] = set()
    ratio_cache: dict[Path, float] = {}
    manifest_entries: list[PlannedOutput] = []

    for row in rows:
        fmt = normalize_format(row.fmt)
        if fmt not in {"svg", "png", "ico"}:
            print(f'SKIP Unsupported format "{row.fmt}" {row_summary(row)}')
            continue

        try:
            path_parts, suggested_path_for_manifest = normalize_suggested_path(row.suggested_path)
        except ValueError as exc:
            print(f"SKIP {exc} {row_summary(row)}")
            continue

        color = normalize_color(row.logo_color)
        if color == "auto":
            expanded_colors = ["light", "dark"]
        elif color in {"light", "dark", "monochrome"}:
            expanded_colors = [color]
        else:
            print(f'SKIP Unsupported Logo Color "{row.logo_color}" {row_summary(row)}')
            continue

        for expanded_color in expanded_colors:
            filename = row.filename
            if color == "auto":
                filename = apply_color_suffix(filename, expanded_color)

            source_svg, source_error = resolve_source_svg(row, expanded_color, filename, fmt, sources_dir)
            if source_error:
                print(f"SKIP {source_error} {row_summary(row, filename)}")
                continue
            if source_svg is None or not source_svg.exists():
                raise RuntimeError(
                    f"Row {row.row_number} failed: source SVG missing ({source_svg}) for "
                    f"{row_summary(row, filename)}"
                )

            variants: list[tuple[str, int | None, int | None, str]] = []
            if fmt == "png":
                export_px_canonical = canonical_export_px(row.export_px)
                if export_px_canonical == "1200x630":
                    print(f"SKIP OG requires template (no og-template provided): {row_summary(row, filename)}")
                    continue

                range_sizes = parse_png_range(row.export_px)
                if range_sizes:
                    for size in range_sizes:
                        variants.append(
                            (insert_suffix(filename, str(size)), size, size, f"{size}×{size}")
                        )
                else:
                    try:
                        width, height = parse_png_dimensions(row.export_px, source_svg, ratio_cache)
                    except ValueError as exc:
                        print(f"SKIP {exc} {row_summary(row, filename)}")
                        continue
                    variants.append((filename, width, height, f"{width}×{height}"))
            elif fmt == "ico":
                try:
                    sizes = parse_ico_sizes(row.export_px)
                except ValueError as exc:
                    print(f"SKIP {exc} {row_summary(row, filename)}")
                    continue
                variants.append((filename, None, None, row.export_px))
            else:
                variants.append((filename, None, None, row.export_px))

            for final_filename_candidate, width, height, export_px_for_manifest in variants:
                rel_output = (
                    path_parts / final_filename_candidate
                    if str(path_parts) != "."
                    else PurePosixPath(final_filename_candidate)
                )
                rel_output = reserve_relative_path(rel_output, used_relative_paths)

                final_filename = rel_output.name
                destination = output_root / Path(rel_output.as_posix())

                planned = PlannedOutput(
                    row_number=row.row_number,
                    platform_use=row.platform_use,
                    logo_type=row.logo_type,
                    logo_color=expanded_color,
                    fmt=fmt,
                    size_spec=row.size_spec,
                    export_px_for_manifest=export_px_for_manifest,
                    suggested_path=suggested_path_for_manifest,
                    final_filename=final_filename,
                    relative_output_path=rel_output,
                    source_svg=source_svg,
                    png_width=width,
                    png_height=height,
                    ico_sizes=parse_ico_sizes(row.export_px) if fmt == "ico" else None,
                )

                if not dry_run:
                    destination.parent.mkdir(parents=True, exist_ok=True)
                    if fmt == "svg":
                        shutil.copyfile(source_svg, destination)
                    elif fmt == "png":
                        if planned.png_width is None or planned.png_height is None:
                            raise RuntimeError(
                                f"Row {row.row_number} failed: PNG dimensions missing for {row_summary(row)}"
                            )
                        cairosvg.svg2png(
                            url=str(source_svg),
                            write_to=str(destination),
                            output_width=planned.png_width,
                            output_height=planned.png_height,
                        )
                    elif fmt == "ico":
                        if not planned.ico_sizes:
                            raise RuntimeError(
                                f"Row {row.row_number} failed: ICO sizes missing for {row_summary(row)}"
                            )
                        render_ico(source_svg, destination, planned.ico_sizes)

                print(f"OK {planned.relative_output_path.as_posix()}")
                manifest_entries.append(planned)

    if not dry_run:
        manifest_path = output_root / "ASSETS_MANIFEST.md"
        write_manifest(manifest_path, manifest_entries, excel_path.name, sheet_name)

    return 0


def main() -> int:
    args = parse_args()
    try:
        return generate_outputs(args)
    except Exception as exc:
        print(f"ERROR {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
